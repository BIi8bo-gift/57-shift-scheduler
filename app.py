import os, json, io
from flask import Flask, request, jsonify, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, static_folder=BASE_DIR)

BATCH_FILLS = [
    PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
    PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
]


def build_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = '排班表'

    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='微软雅黑', bold=True, size=14)
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    data_font = Font(name='微软雅黑', size=10)
    note_font = Font(name='微软雅黑', size=9, italic=True, color='555555')
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    batchnames = data.get('batchNames', ['批号1', '批号2', '批号3'])
    headers = ['日期', batchnames[0], batchnames[1], batchnames[2],
               '早班人员', '小夜班人员', '大夜班人员', '休班人员']
    ncols = len(headers)
    last_letter = get_column_letter(ncols)

    col_widths = [16, 36, 36, 36, 30, 30, 30, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f'A1:{last_letter}1')
    c = ws['A1']
    c.value = data.get('batchLabel', '排班表')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f'A2:{last_letter}2')
    c = ws['A2']
    dr = data.get('dateRange', '')
    c.value = '周期: ' + dr if dr else ''
    c.font = note_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border_all
    ws.row_dimensions[3].height = 30

    rows = data.get('rows', [])
    for idx, r in enumerate(rows):
        rn = 4 + idx
        vals = [r.get(k, '') for k in
                ['date', 'batch1', 'batch2', 'batch3', 'morning', 'evening', 'night', 'off']]
        max_lines = 1
        for v in vals:
            lines = str(v).count('\n') + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[rn].height = max(72, max_lines * 18)

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=col, value=val)
            cell.font = data_font
            cell.border = border_all
            cell.alignment = center if col == 1 else left_wrap
            if 2 <= col <= 4:
                cell.fill = BATCH_FILLS[col - 2]

    sr = 4 + len(rows)
    ws.merge_cells(f'A{sr}:D{sr}')
    c = ws.cell(row=sr, column=1, value='合计')
    c.font = Font(name='微软雅黑', bold=True, size=10)
    c.alignment = center
    c.border = border_all
    for col in range(2, ncols + 1):
        ws.cell(row=sr, column=col).border = border_all
    ws.cell(row=sr, column=5, value=data.get('summaryText', '')).font = Font(name='微软雅黑', size=10)
    ws.cell(row=sr, column=5).alignment = left_wrap
    ws.cell(row=sr, column=5).border = border_all
    ws.row_dimensions[sr].height = 25

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')


@app.route('/export', methods=['POST'])
def export():
    payload = request.get_json()
    if not payload:
        return jsonify({'error': 'no data'}), 400
    try:
        buf = build_xlsx(payload)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='57车间排班表.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===== 服务端保存/加载 =====

SAVE_FILE = os.path.join(BASE_DIR, 'saved_data.json')


@app.route('/api/save', methods=['POST'])
def api_save():
    payload = request.get_json()
    if not payload:
        return jsonify({'error': 'no data'}), 400
    try:
        with open(SAVE_FILE, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/load', methods=['GET'])
def api_load():
    try:
        if os.path.exists(SAVE_FILE):
            with open(SAVE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify(data)
        return jsonify(None)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===== PDF Export =====

class SchedulePDF(FPDF):
    FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts')

    def __init__(self):
        super().__init__('L', 'mm', 'A4')
        self.set_auto_page_break(auto=False)
        font_dir = self.FONT_DIR
        self.add_font('NotoSC', '', os.path.join(font_dir, 'NotoSansSC-Regular-subset.ttf'))
        self.add_font('NotoSC', 'B', os.path.join(font_dir, 'NotoSansSC-Bold-subset.ttf'))

    def footer(self):
        self.set_y(-12)
        self.set_font('NotoSC', '', 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, '57车间排班表 - 第 %d/{nb} 页' % self.page_no(), align='C')


def build_pdf(data):
    pdf = SchedulePDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    HDR_BG = (68, 114, 196)
    HDR_FG = (255, 255, 255)
    SUM_BG = (240, 240, 240)
    BLACK = (30, 30, 30)
    GRAY = (120, 120, 120)

    # 获取班次标题（用户可能自定义过）
    shift_keys = ['e', 'mb', 'mp', 's1', 's2', 'bn', 'of']
    shift_headers = data.get('shiftHeaders', ['早班','中午值班','延迟吃饭','小夜1:30','小夜3:30','大夜','休班'])
    headers = ['日期'] + shift_headers  # 8列
    ncols = len(headers)

    # A4横向297mm，左右边距各留空
    col_w = [22] + [28] * 7  # sum=22+196=218mm
    total_w = sum(col_w)
    margin_left = (297 - total_w) / 2

    title = data.get('batchLabel', '57车间排班表')
    date_range = data.get('dateRange', '')

    content_y = 15
    max_y = 265
    LINE_H = 4.8
    FONT_SZ = 7
    HDR_H = 8

    # --- Title ---
    pdf.set_xy(margin_left, content_y)
    pdf.set_font('NotoSC', 'B', 14)
    pdf.set_text_color(*BLACK)
    pdf.cell(total_w, 10, title, align='C')
    content_y += 12

    # --- Date range ---
    pdf.set_xy(margin_left, content_y)
    pdf.set_font('NotoSC', '', 8)
    pdf.set_text_color(*GRAY)
    pdf.cell(total_w, 6, '周期: ' + date_range, align='C')
    content_y += 9

    def draw_header(y):
        x = margin_left
        pdf.set_font('NotoSC', 'B', FONT_SZ)
        pdf.set_fill_color(*HDR_BG)
        pdf.set_text_color(*HDR_FG)
        for i, h in enumerate(headers):
            pdf.set_xy(x, y)
            pdf.cell(col_w[i], HDR_H, h, border=1, fill=True, align='C')
            x += col_w[i]
        return y + HDR_H + 1

    def draw_data_row(y, vals):
        """先渲染文字确定实际高度，再画边框"""
        bottoms = [y + 8]
        for col_i, val in enumerate(vals):
            display = str(val)
            if not display:
                continue
            cx = margin_left + sum(col_w[:col_i])
            pdf.set_xy(cx + 0.5, y + 0.5)
            pdf.set_font('NotoSC', '', FONT_SZ)
            pdf.set_text_color(*BLACK)
            pdf.multi_cell(col_w[col_i] - 1, LINE_H, display, align='L')
            bottoms.append(pdf.get_y())

        actual_bottom = max(bottoms)
        row_h = actual_bottom - y

        pdf.set_draw_color(100, 100, 100)
        for col_i in range(ncols):
            cx = margin_left + sum(col_w[:col_i])
            pdf.rect(cx, y, col_w[col_i], row_h)

        return actual_bottom

    def draw_summary_row(y, summary_text=''):
        x = margin_left
        pdf.set_font('NotoSC', 'B', FONT_SZ)
        pdf.set_fill_color(*SUM_BG)
        pdf.set_text_color(*BLACK)
        pdf.set_xy(x, y)
        pdf.cell(col_w[0], HDR_H, '合计', border=1, fill=True, align='C')
        x += col_w[0]
        sum_w = sum(col_w[1:])
        pdf.set_font('NotoSC', '', FONT_SZ)
        pdf.set_xy(x, y)
        display = str(summary_text)[:50] if summary_text else ''
        pdf.cell(sum_w, HDR_H, display, border=1, fill=True, align='C')
        return y + HDR_H

    # Draw header
    content_y = draw_header(content_y)

    rows = data.get('rows', [])
    for idx, r in enumerate(rows):
        vals = [r.get(k, '') for k in ['date'] + shift_keys]

        # 粗略估算行高用于分页判断
        est_h = 8
        for v in vals:
            t = str(v)
            lines = max(1, (len(t) // 10) + 1)
            est_h = max(est_h, lines * LINE_H + 2)
        est_h += 3

        if content_y + est_h > max_y:
            pdf.add_page()
            content_y = 15
            content_y = draw_header(content_y)

        content_y = draw_data_row(content_y, vals)

    # Summary row
    content_y += 3
    if content_y + HDR_H > max_y:
        pdf.add_page()
        content_y = 15
    draw_summary_row(content_y, data.get('summaryText', ''))

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


@app.route('/export-pdf', methods=['POST'])
def export_pdf():
    payload = request.get_json()
    if not payload:
        return jsonify({'error': 'no data'}), 400
    try:
        buf = build_pdf(payload)
        return send_file(
            buf,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='57车间排班表.pdf'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=True)
