import http.server, socketserver, json, os, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

WEBROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)))


BATCH_FILLS = [
    PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),  # green
    PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),  # blue
    PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),  # yellow
]


def build_xlsx(data):
    """Generate formatted xlsx.
    8-column format: 日期 | batch1 | batch2 | batch3 | 早班人员 | 小夜班人员 | 大夜班人员 | 休班人员
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '排班表'

    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='微软雅黑', bold=True, size=14)
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    data_font = Font(name='微软雅黑', size=10)
    note_font = Font(name='微软雅黑', size=9, italic=True, color='555555')
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Columns: 日期 | batch1 | batch2 | batch3 | 早班 | 小夜班 | 大夜班 | 休班
    col_widths = [16, 36, 36, 36, 30, 30, 30, 16]
    batch_names = data.get('batchNames', ['批号1', '批号2', '批号3'])
    headers = ['日期', batch_names[0], batch_names[1], batch_names[2],
               '早班人员', '小夜班人员', '大夜班人员', '休班人员']
    ncols = len(headers)  # 8
    last_letter = get_column_letter(ncols)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: Title
    ws.merge_cells(f'A1:{last_letter}1')
    c = ws['A1']
    c.value = data.get('batchLabel', '排班表')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    # Row 2: Info bar
    ws.merge_cells(f'A2:{last_letter}2')
    c = ws['A2']
    date_range = data.get('dateRange', '')
    c.value = '周期: ' + date_range if date_range else ''
    c.font = note_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    # Row 3: Table header
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border_all
    ws.row_dimensions[3].height = 30

    # Data rows
    rows = data.get('rows', [])
    for idx, r in enumerate(rows):
        row_num = 4 + idx
        vals = [
            r.get('date', ''),
            r.get('batch1', ''),
            r.get('batch2', ''),
            r.get('batch3', ''),
            r.get('morning', ''),
            r.get('evening', ''),
            r.get('night', ''),
            r.get('off', ''),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.border = border_all
            cell.alignment = center if col == 1 else left_wrap
            # Batch columns get colored backgrounds
            if 2 <= col <= 4:
                cell.fill = BATCH_FILLS[col - 2]
        ws.row_dimensions[row_num].height = 72

    # Summary row
    sr = 4 + len(rows)
    ws.merge_cells(f'A{sr}:D{sr}')
    c = ws.cell(row=sr, column=1, value='合计')
    c.font = Font(name='微软雅黑', bold=True, size=10)
    c.alignment = center
    c.border = border_all
    for col in range(2, ncols + 1):
        ws.cell(row=sr, column=col).border = border_all
    ws.cell(row=sr, column=5, value=data.get('summaryText', '')).font = Font(name='微软雅黑', size=10)
    ws.cell(row=sr, column=5).alignment = left_wrap
    ws.cell(row=sr, column=5).border = border_all
    ws.row_dimensions[sr].height = 25

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        # Serve static files from WEBROOT
        if self.path == '/':
            self.path = '/index.html'
        # Let parent class handle the GET
        return http.server.SimpleHTTPRequestHandler.do_GET(self)

    def do_POST(self):
        if self.path == '/export':
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            try:
                payload = json.loads(body)
            except Exception:
                self.send_error(400, 'Invalid JSON')
                return
            try:
                xlsx_buf = build_xlsx(payload)
                self.send_response(200)
                self.send_header('Content-Type',
                                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition',
                                 'attachment; filename="shift-schedule.xlsx"')
                self.send_header('Content-Length', str(xlsx_buf.getbuffer().nbytes))
                self.end_headers()
                self.wfile.write(xlsx_buf.getvalue())
            except Exception as e:
                self.send_error(500, str(e))
        else:
            self.send_error(404)


if __name__ == '__main__':
    os.chdir(WEBROOT)
    httpd = socketserver.TCPServer(('0.0.0.0', 8080), Handler)
    print(f'Serving http://127.0.0.1:8080/ from {WEBROOT}')
    httpd.serve_forever()
